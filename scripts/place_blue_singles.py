#!/usr/bin/env python3
"""
Auto-place remaining "blue" single categories into parent groups using a scored
match against the whole "Regrouped" sheet. 10 rounds; each round only re-scores
singles still unplaced, so as groups grow, later rounds catch more.

Deterministic. No fuzzy/embedding/LLM matching — pure token scoring, auditable.
Placements are PROPOSALS (blue) only; nothing is pushed to the DB.

IN : ~/Desktop/sami_v9.xlsx  (sheet "Regrouped")
OUT: ~/Desktop/sami_v10.xlsx (sheet "Regrouped") + ~/Desktop/placements.csv
"""
import csv
import os
import re
from copy import copy
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font

DESKTOP = os.path.expanduser("~/Desktop")
SRC = os.path.join(DESKTOP, "sami_v9.xlsx")
OUT_XLSX = os.path.join(DESKTOP, "sami_v10.xlsx")
OUT_CSV = os.path.join(DESKTOP, "placements.csv")

BLUE = "FF00B0F0"
RED_TEXT = "FFFF0000"
MEMBER_PREFIX = "        ▶ (placed) "   # matches existing ▶ placed copies (8 spaces)

# ── tokenizer / stemmer (exactly per spec) ────────────────────────────────────
SUFFIXES = ["ing", "ers", "er", "ors", "or", "ions", "ion", "ies", "es", "s"]

def stem(word):
    w = word.lower()
    for suf in SUFFIXES:
        if w.endswith(suf) and len(w) - len(suf) >= 3:
            return w[:-len(suf)]
    return w

# Generic stopwords, with the "(s)" variants expanded to concrete surface forms.
STOP = {
    "and", "the", "of", "general", "pty", "australia", "group", "new",
    "service", "services", "solution", "solutions", "system", "systems",
    "specialist", "specialists", "company", "contractor", "contractors",
    "commercial", "residential", "domestic", "installation", "install",
    "repair", "repairs", "supplier", "suppliers", "supply",
    "product", "products", "equipment", "maintenance", "work", "works",
    "building",
}

def tokens(text):
    if not text:
        return []
    t = str(text).lower().replace("&", " and ")
    t = re.sub(r"[^a-z0-9]+", " ", t)
    out = []
    for raw in t.split():
        if len(raw) < 1 or raw in STOP:      # drop stopwords on surface form
            continue
        s = stem(raw)
        if len(s) < 3 or s in STOP:          # drop short + stopword stems
            continue
        out.append(s)
    return out

def primary(fullname):
    """Tokens of the FIRST '/'-separated segment only (the parent tag)."""
    first = str(fullname or "").split("/")[0]
    return tokens(first)

def leaf(fullname):
    return str(fullname or "").split("/")[0].strip()

# ── load workbook (with styles) ───────────────────────────────────────────────
wb = openpyxl.load_workbook(SRC)
ws = wb["Regrouped"]
MAXC = 6
n_rows = ws.max_row

def kind_of(c1):
    s = str(c1 or "")
    if s.startswith("MERGED"):
        return "header"
    if "↳" in s:
        return "member"
    if "▶" in s:
        return "member"      # already-placed copies are part of the group now
    if s.strip() == "":
        return "blank"
    return "single"

# Snapshot every original row: values + a *copy* of each cell's style.
orig = []   # list of dict{vals:[6], styles:[6], kind, level, cid, slug, full, status}
for r in range(1, n_rows + 1):
    vals, styles = [], []
    for c in range(1, MAXC + 1):
        cell = ws.cell(r, c)
        vals.append(cell.value)
        styles.append(copy(cell._style))
    orig.append({
        "vals": vals, "styles": styles, "kind": kind_of(vals[0]),
        "level": vals[1], "cid": vals[2], "full": vals[3],
        "slug": vals[4], "status": vals[5],
    })

# Capture style templates BEFORE we overwrite anything.
# row 23 = an existing ▶ placed member (blue); row 1419 = a blue placed single.
BLUE_MEMBER_STYLES = [copy(ws.cell(23, c)._style) for c in range(1, MAXC + 1)]
BLUE_SINGLE_STYLES = [copy(ws.cell(1419, c)._style) for c in range(1, MAXC + 1)]

# ── parse groups (header row 0-based index -> group) ──────────────────────────
groups = {}          # gid -> group dict
cur = None
for i, row in enumerate(orig):
    if i == 0:
        continue     # header row of the sheet
    k = row["kind"]
    if k == "header":
        name = str(row["vals"][0]).split("→", 1)[1].strip()
        cur = i
        groups[cur] = {
            "gid": cur, "name": name, "header_idx": i,
            "last_member_idx": i,          # updated as members are seen
            "members": [],                  # list of dict{level,cid,slug,full}
            "added": [],                    # fullnames placed this run
        }
    elif k == "member" and cur is not None:
        groups[cur]["members"].append({
            "level": row["level"], "cid": row["cid"],
            "slug": row["slug"], "full": row["full"],
        })
        groups[cur]["last_member_idx"] = i
    elif k == "single":
        cur = None    # entered the singles list; singles never join a group

# Blue singles to place = single rows with status exactly "no match".
blue_singles = [i for i, row in enumerate(orig)
                if row["kind"] == "single" and row["status"] == "no match"]

print(f"Groups: {len(groups)}   Blue singles (no match): {len(blue_singles)}")

# ── vocab builders (recomputed each round because groups grow) ─────────────────
def build_vocab():
    label_roots, prim_roots, name_roots, size = {}, {}, {}, {}
    group_count = defaultdict(int)          # token -> #distinct groups it appears in
    for gid, g in groups.items():
        fulls = [m["full"] for m in g["members"]] + g["added"]
        lab = set(tokens(g["name"]))
        prm = set()
        nam = set()
        for f in fulls:
            prm |= set(primary(f))
            nam |= set(tokens(f))
        label_roots[gid] = lab
        prim_roots[gid] = prm
        name_roots[gid] = nam
        size[gid] = len(fulls)
        for t in nam:
            group_count[t] += 1
    common = {t for t, c in group_count.items() if c > 10}
    return label_roots, prim_roots, name_roots, size, common

def score_single(full, label_roots, prim_roots, name_roots, common):
    sp = set(primary(full))
    sa = set(tokens(full))
    ranked = []
    for gid, g in groups.items():
        sc = 3 if (sp & (label_roots[gid] | prim_roots[gid])) else 0
        sc += sum(1 for t in sa if t in name_roots[gid] and t not in common)
        ranked.append((sc, gid))
    return ranked, sa

def survivor(g):
    """Winning group's survivor = lowest-Level, lowest-ID member."""
    ms = [m for m in g["members"] if m["cid"] is not None]
    if not ms:
        return (None, None)
    m = min(ms, key=lambda m: ((m["level"] if m["level"] is not None else 999), m["cid"]))
    return (m["cid"], m["slug"])

# ── 10-round placement loop ───────────────────────────────────────────────────
placement = {}       # single orig-index -> dict(gid, score, round, alts, ...)
remaining = list(blue_singles)
cumulative = 0
print("\nround  placed  cumulative  remaining")
for rnd in range(1, 11):
    label_roots, prim_roots, name_roots, size, common = build_vocab()
    placed_this = []
    for si in remaining:
        full = orig[si]["full"]
        ranked, _ = score_single(full, label_roots, prim_roots, name_roots, common)
        # sort: score desc, size desc, name asc  (deterministic)
        ranked.sort(key=lambda x: (-x[0], -size[x[1]], groups[x[1]]["name"]))
        best_sc, best_gid = ranked[0]
        if best_sc >= 3:
            # did the parent-name (+3) signal fire for the winner?
            sp = set(primary(full))
            parent = bool(sp & (label_roots[best_gid] | prim_roots[best_gid]))
            low_conf = (not parent) or (best_sc == 3)
            alts = [(groups[g]["name"], s) for s, g in ranked[1:3]]
            placement[si] = {
                "gid": best_gid, "score": best_sc, "round": rnd, "alts": alts,
                "parent": parent, "low_conf": low_conf,
            }
            placed_this.append((si, best_gid, full))

    # apply placements: grow the winning groups so later rounds see them
    for si, gid, full in placed_this:
        groups[gid]["added"].append(full)
    for si, _, _ in placed_this:
        remaining.remove(si)
    cumulative += len(placed_this)
    print(f"{rnd:>5}  {len(placed_this):>6}  {cumulative:>10}  {len(remaining):>9}")
    if not placed_this:
        break

# ── build output row list (originals + inserted blue group-copies) ────────────
# queue copies per group header index
copies_by_group = defaultdict(list)     # gid -> list of copy-row descriptors
for si, p in placement.items():
    row = orig[si]
    gid = p["gid"]
    tag = " LOW-CONF" if p["low_conf"] else ""
    copy_vals = [
        MEMBER_PREFIX + leaf(row["full"]),
        row["level"], row["cid"], row["full"], row["slug"],
        f"PLACED r{p['round']} score {p['score']}{tag}",
    ]
    copies_by_group[gid].append({
        "sort": (p["round"], p["score"], str(row["full"])),
        "vals": copy_vals,
        "styles": [copy(s) for s in BLUE_MEMBER_STYLES],
        "low_conf": p["low_conf"],
    })
for gid in copies_by_group:
    copies_by_group[gid].sort(key=lambda d: d["sort"])

# recolor placed singles in-place + set status; leave unplaced as "no match".
for si, p in placement.items():
    gname = groups[p["gid"]]["name"]
    alts = p["alts"]
    alt_str = " | ".join(f"{name}({s})" for name, s in alts)
    tag = " · LOW-CONF" if p["low_conf"] else ""
    orig[si]["vals"][5] = f"→ PLACED: {gname} ({p['score']}) · alts: {alt_str}{tag}"
    orig[si]["styles"] = [copy(s) for s in BLUE_SINGLE_STYLES]

# assemble final ordered rows; track which FINAL rows are low-confidence (red text)
out_rows = []
low_conf_rows = []                       # 1-based final row numbers to paint red
last_member_to_gid = {g["last_member_idx"]: gid for gid, g in groups.items()}
for i, row in enumerate(orig):
    out_rows.append({"vals": row["vals"], "styles": row["styles"]})
    if i in placement and placement[i]["low_conf"]:
        low_conf_rows.append(len(out_rows))          # the placed single row
    gid = last_member_to_gid.get(i)
    if gid is not None:
        for cp in copies_by_group.get(gid, []):
            out_rows.append({"vals": cp["vals"], "styles": cp["styles"]})
            if cp["low_conf"]:
                low_conf_rows.append(len(out_rows))  # the ▶ copy row

# ── write back into the Regrouped sheet ───────────────────────────────────────
for idx, r in enumerate(out_rows, start=1):
    for c in range(1, MAXC + 1):
        cell = ws.cell(idx, c)
        cell.value = r["vals"][c - 1]
        cell._style = r["styles"][c - 1]

# paint low-confidence placement rows with RED text (blue fill is preserved)
for idx in low_conf_rows:
    for c in range(1, MAXC + 1):
        cell = ws.cell(idx, c)
        f = cell.font
        cell.font = Font(name=f.name, size=f.size, bold=f.bold,
                         italic=f.italic, underline=f.underline, color=RED_TEXT)

# ── assertions: no category lost or invented ──────────────────────────────────
orig_ids = {row["cid"] for row in orig if row["cid"] is not None}
new_ids = defaultdict(int)
for r in out_rows:
    cid = r["vals"][2]
    if cid is not None:
        new_ids[cid] += 1
new_id_set = set(new_ids)
if new_id_set != orig_ids:
    missing = orig_ids - new_id_set
    invented = new_id_set - orig_ids
    raise SystemExit(f"ASSERT FAILED: category id set changed. missing={missing} invented={invented}")
n_copies = sum(len(v) for v in copies_by_group.values())
if len(out_rows) != len(orig) + n_copies:
    raise SystemExit(f"ASSERT FAILED: row math off. out={len(out_rows)} orig={len(orig)} copies={n_copies}")
print(f"\nASSERT OK: distinct real categories unchanged = {len(orig_ids)} "
      f"(rows {len(orig)} -> {len(out_rows)}, +{n_copies} blue copies)")

# ── placements.csv ────────────────────────────────────────────────────────────
with open(OUT_CSV, "w", newline="") as fh:
    w = csv.writer(fh)
    w.writerow(["old_id", "old_slug", "target_label", "target_id",
                "target_slug", "round", "score", "confidence", "parent_match"])
    for si, p in sorted(placement.items(),
                        key=lambda kv: (kv[1]["low_conf"], kv[1]["round"], -kv[1]["score"])):
        row = orig[si]
        tid, tslug = survivor(groups[p["gid"]])
        w.writerow([row["cid"], row["slug"], groups[p["gid"]]["name"],
                    tid, tslug, p["round"], p["score"],
                    "low" if p["low_conf"] else "high",
                    "yes" if p["parent"] else "no"])

wb.save(OUT_XLSX)
n_low = sum(1 for p in placement.values() if p["low_conf"])
n_high = len(placement) - n_low
print(f"\nConfidence: {n_high} high, {n_low} low (red text)")
print(f"\nWrote {OUT_XLSX}")
print(f"Wrote {OUT_CSV}  ({len(placement)} placements, {len(remaining)} still unplaced)")
