#!/usr/bin/env python3
"""
RE-SCORE every blue proposal + no-match single against the TRUSTED group
structure and RE-HOME each to its best parent. Supersedes place_blue_singles.py.

Two corrections over the previous run:
  1. Parent-name match reads the WHOLE DB category name (all "/" segments), not
     just the first — so e.g. "Ground Engineering / Resin Injection / Underpinning"
     matches the *Underpinning* parent via its 3rd segment.
  2. A group's matching vocabulary comes ONLY from TRUSTED members
     (status "moved (manual)" or "merged"). Prior blue proposals never vote, so
     earlier mistakes can't snowball.

Re-home: all existing "▶ (placed)" copies are dropped and regenerated fresh under
each category's new winning group. Trusted members + headers are never touched.

Deterministic. Proposals only (blue); nothing is pushed to the DB.

IN : ~/Desktop/sami_v9.xlsx  (sheet "Regrouped")
OUT: ~/Desktop/sami_v10.xlsx (sheet "Regrouped") + ~/Desktop/placements.csv
"""
import csv
import os
import re
from copy import copy
from collections import defaultdict, Counter

import openpyxl
from openpyxl.styles import Font

DESKTOP = os.path.expanduser("~/Desktop")
SRC = os.path.join(DESKTOP, "sami_v9.xlsx")
OUT_XLSX = os.path.join(DESKTOP, "sami_v10.xlsx")
OUT_CSV = os.path.join(DESKTOP, "placements.csv")

BLUE = "FF00B0F0"
RED_TEXT = "FFFF0000"
MEMBER_PREFIX = "        ▶ (placed) "
TRUSTED_STATUS = {"moved (manual)", "merged"}

# ── Explicit, auditable synonym groups (stemmed forms) ────────────────────────
# A blue-category word counts as a "synonym match" to a parent keyword when both
# sit in the same group below. At most ONE synonym may count toward a placement
# (the 3rd match); the other 2 must be exact keyword hits. Edit freely.
SYNONYM_GROUPS = [
    {"waterproof", "tank", "membran"},
    {"underpin", "foundat", "restump", "footing"},
    {"damp", "moistur", "moisture"},
    {"crack", "fractur"},
    {"seal", "sealant", "caulk"},
    {"render", "plaster"},
    {"concret", "cement", "cementiti"},
    {"gutter", "gutt", "downpip", "fascia"},
    {"glaz", "glass", "window"},
    {"facad", "clad", "cladd"},
    {"til", "tile", "tiling"},
    {"paint", "coat", "coating"},
    {"drain", "stormwat", "storm"},
    {"insul", "insulat"},
    {"asphalt", "bitumen"},
    {"demolit", "strip", "excavat"},
]
SYN = {}
for _gi, _grp in enumerate(SYNONYM_GROUPS):
    for _w in _grp:
        SYN[_w] = _gi        # word -> synonym-group id

# ── tokenizer / stemmer (unchanged, per original spec) ────────────────────────
SUFFIXES = ["ing", "ers", "er", "ors", "or", "ions", "ion", "ies", "es", "s"]

def stem(word):
    w = word.lower()
    for suf in SUFFIXES:
        if w.endswith(suf) and len(w) - len(suf) >= 3:
            return w[:-len(suf)]
    return w

STOP = {
    "and", "the", "of", "general", "pty", "australia", "group", "new",
    "service", "services", "solution", "solutions", "system", "systems",
    "specialist", "specialists", "company", "contractor", "contractors",
    "commercial", "residential", "domestic", "installation", "install",
    "repair", "repairs", "supplier", "suppliers", "supply",
    "product", "products", "equipment", "maintenance", "work", "works",
    "building",
}

def tokens(text):
    if not text:
        return []
    t = str(text).lower().replace("&", " and ")
    t = re.sub(r"[^a-z0-9]+", " ", t)
    out = []
    for raw in t.split():
        if raw in STOP:
            continue
        s = stem(raw)
        if len(s) < 3 or s in STOP:
            continue
        out.append(s)
    return out

def leaf(fullname):
    return str(fullname or "").split("/")[0].strip()

# ── load workbook + snapshot rows/styles ──────────────────────────────────────
wb = openpyxl.load_workbook(SRC)
ws = wb["Regrouped"]
MAXC = 6
n_rows = ws.max_row

def kind_of(c1, status, is_blue):
    s = str(c1 or "")
    if s.startswith("MERGED"):
        return "header"
    if "↳" in s:
        return "trusted"        # all ↳ members are moved/merged (verified)
    if "▶" in s:
        return "arrow"          # prior proposal copy — drop + regenerate
    # single row in the list
    if is_blue or status == "no match":
        return "rescore"        # blue proposal or unplaced → re-score
    return "keep"               # cross-listed / other — leave untouched

def is_blue_cell(r):
    f = ws.cell(r, 1).fill
    return bool(f.patternType) and f.fgColor.rgb == BLUE

orig = []
for r in range(1, n_rows + 1):
    vals = [ws.cell(r, c).value for c in range(1, MAXC + 1)]
    styles = [copy(ws.cell(r, c)._style) for c in range(1, MAXC + 1)]
    k = kind_of(vals[0], vals[5], is_blue_cell(r)) if r > 1 else "sheethdr"
    orig.append({"vals": vals, "styles": styles, "kind": k,
                 "level": vals[1], "cid": vals[2], "full": vals[3],
                 "slug": vals[4], "status": vals[5]})

BLUE_MEMBER_STYLES = [copy(ws.cell(23, c)._style) for c in range(1, MAXC + 1)]
BLUE_SINGLE_STYLES = [copy(ws.cell(1419, c)._style) for c in range(1, MAXC + 1)]

# ── parse groups: trusted-only vocabulary + prior-parent map ──────────────────
groups = {}                 # gid (header idx) -> group dict
prev_parent = {}            # category id -> name of group it was a ▶ copy in
cur = None
for i, row in enumerate(orig):
    if i == 0:
        continue
    k = row["kind"]
    if k == "header":
        name = str(row["vals"][0]).split("→", 1)[1].strip()
        cur = i
        groups[cur] = {"gid": cur, "name": name, "header_idx": i,
                       "trusted": [], "added": []}
    elif k == "trusted" and cur is not None:
        groups[cur]["trusted"].append(
            {"level": row["level"], "cid": row["cid"],
             "slug": row["slug"], "full": row["full"]})
    elif k == "arrow" and cur is not None:
        prev_parent[row["cid"]] = groups[cur]["name"]
    elif k in ("rescore", "keep"):
        cur = None

rescore = [i for i, row in enumerate(orig) if row["kind"] == "rescore"]
print(f"Groups: {len(groups)}   Trusted members: {sum(len(g['trusted']) for g in groups.values())}")
print(f"Categories to re-score (blue + no-match): {len(rescore)}")

# ── vocab builders (trusted + this-run placements) ────────────────────────────
def build_keywords():
    # keywords(parent) = up to 3 words with the highest document-frequency across
    # the parent's TRUSTED members (+ this-run placements). Only genuinely repeated
    # words (df >= 2) qualify — "most repeated word between trusted categories".
    keywords, size = {}, {}
    for gid, g in groups.items():
        fulls = [m["full"] for m in g["trusted"]] + g["added"]
        size[gid] = len(fulls)
        df = Counter()
        for f in fulls:
            for w in set(tokens(f)):               # per-member distinct = doc freq
                df[w] += 1
        repeated = [(w, c) for w, c in df.items() if c >= 2]
        repeated.sort(key=lambda wc: (-wc[1], wc[0]))   # freq desc, then alpha
        keywords[gid] = [w for w, _ in repeated[:3]]     # Option 4: top-3 keywords
    return keywords, size

def score_all(full, keywords):
    # score = exact keyword hits + at most ONE synonym hit (the allowed 3rd match).
    sa = set(tokens(full))
    ranked = []
    for gid in groups:
        kws = keywords[gid]
        kwset = set(kws)
        exact = len(sa & kwset)
        kw_groups = {SYN[k] for k in kws if k in SYN}
        syn = sum(1 for w in sa
                  if w not in kwset and SYN.get(w) in kw_groups and w in SYN)
        score = exact + min(1, syn)
        ranked.append((score, gid, exact, min(1, syn)))
    return ranked

def survivor(g):
    ms = [m for m in g["trusted"] if m["cid"] is not None]
    if not ms:
        return (None, None)
    m = min(ms, key=lambda m: ((m["level"] if m["level"] is not None else 999), m["cid"]))
    return (m["cid"], m["slug"])

# ── 10-round loop (base vocab = trusted; new placements grow their group) ──────
placement = {}
remaining = list(rescore)
cumulative = 0
print("\nround  placed  cumulative  remaining")
for rnd in range(1, 11):
    keywords, size = build_keywords()
    placed_this = []
    for si in remaining:
        full = orig[si]["full"]
        ranked = score_all(full, keywords)
        ranked.sort(key=lambda x: (-x[0], -size[x[1]], groups[x[1]]["name"]))
        best_sc, best_gid, best_exact, best_syn = ranked[0]
        if best_sc >= 2:                            # Option 4: min 2 (2 kw, or 1 kw + 1 synonym)
            alts = [(groups[g]["name"], s) for s, g, _, _ in ranked[1:3]]
            placement[si] = {
                "gid": best_gid, "score": best_sc, "round": rnd, "alts": alts,
                "parent": (best_exact >= 3),
                "low_conf": (best_sc < 3) or (best_syn >= 1),  # high only = pure 3-exact
                "exact": best_exact, "syn": best_syn,
            }
            placed_this.append((si, best_gid, full))
    for si, gid, full in placed_this:
        groups[gid]["added"].append(full)
    for si, _, _ in placed_this:
        remaining.remove(si)
    cumulative += len(placed_this)
    print(f"{rnd:>5}  {len(placed_this):>6}  {cumulative:>10}  {len(remaining):>9}")
    if not placed_this:
        break

# ── build fresh ▶ copies + recolor singles-list rows ──────────────────────────
copies_by_group = defaultdict(list)
for si, p in placement.items():
    row = orig[si]
    tag = " LOW-CONF" if p["low_conf"] else ""
    copies_by_group[p["gid"]].append({
        "sort": (p["round"], p["score"], str(row["full"])),
        "vals": [MEMBER_PREFIX + leaf(row["full"]), row["level"], row["cid"],
                 row["full"], row["slug"], f"PLACED r{p['round']} score {p['score']}{tag}"],
        "styles": [copy(s) for s in BLUE_MEMBER_STYLES],
        "low_conf": p["low_conf"],
    })
for gid in copies_by_group:
    copies_by_group[gid].sort(key=lambda d: d["sort"])

# recolor placed singles; revert now-unplaced ones back to plain "no match".
placed_set = set(placement)
for si in rescore:
    row = orig[si]
    if si in placed_set:
        p = placement[si]
        gname = groups[p["gid"]]["name"]
        alt_str = " | ".join(f"{n}({s})" for n, s in p["alts"])
        tag = " · LOW-CONF" if p["low_conf"] else ""
        row["vals"][5] = f"→ PLACED: {gname} ({p['score']}) · alts: {alt_str}{tag}"
        row["styles"] = [copy(s) for s in BLUE_SINGLE_STYLES]
    else:
        # unplaced now: plain white, "no match"
        row["vals"][5] = "no match"
        row["styles"] = [copy(orig[0]["styles"][c]) for c in range(MAXC)]  # header row = plain default

# ── assemble output: drop arrows, keep everything else, splice fresh copies ────
base = []                       # list of {vals, styles}
pos2gid = {}                    # base index after which to insert a group's copies
cur_gid = None
for i, row in enumerate(orig):
    k = row["kind"]
    if k == "arrow":
        continue                # drop prior proposal copies
    base.append({"vals": row["vals"], "styles": row["styles"]})
    if k == "header":
        cur_gid = i
        pos2gid[len(base) - 1] = cur_gid           # default insert after header
    elif k == "trusted":
        pos2gid[len(base) - 1] = cur_gid           # push to last trusted member
    elif k in ("rescore", "keep"):
        cur_gid = None

# keep only the LAST recorded position per gid
last_pos = {}
for pos, gid in pos2gid.items():
    if gid is None:
        continue
    if gid not in last_pos or pos > last_pos[gid]:
        last_pos[gid] = pos
insert_at = {pos: gid for gid, pos in last_pos.items()}

final = []
low_conf_rows = []
for idx, rd in enumerate(base):
    final.append(rd)
    gid = insert_at.get(idx)
    if gid is not None:
        for cp in copies_by_group.get(gid, []):
            final.append({"vals": cp["vals"], "styles": cp["styles"]})
# low-conf rows (both ▶ copies and singles-list rows) carry a LOW-CONF marker
low_conf_rows = [ridx for ridx, rd in enumerate(final, start=1)
                 if "LOW-CONF" in str(rd["vals"][5] or "")]

# ── write into the sheet ──────────────────────────────────────────────────────
for idx, rd in enumerate(final, start=1):
    for c in range(1, MAXC + 1):
        cell = ws.cell(idx, c)
        cell.value = rd["vals"][c - 1]
        cell._style = rd["styles"][c - 1]
# clear any trailing rows left over from the (larger) original sheet
if len(final) < n_rows:
    for r in range(len(final) + 1, n_rows + 1):
        for c in range(1, MAXC + 1):
            ws.cell(r, c).value = None

for idx in low_conf_rows:
    for c in range(1, MAXC + 1):
        cell = ws.cell(idx, c)
        f = cell.font
        cell.font = Font(name=f.name, size=f.size, bold=f.bold,
                         italic=f.italic, underline=f.underline, color=RED_TEXT)

# ── assertions: no category lost or invented ──────────────────────────────────
orig_ids = {row["cid"] for row in orig if row["cid"] is not None}
new_ids = {rd["vals"][2] for rd in final if rd["vals"][2] is not None}
if new_ids != orig_ids:
    raise SystemExit(f"ASSERT FAILED: id set changed. "
                     f"missing={orig_ids - new_ids} invented={new_ids - orig_ids}")
n_copies = sum(len(v) for v in copies_by_group.values())
print(f"\nASSERT OK: distinct real categories unchanged = {len(orig_ids)} "
      f"(rows {n_rows} -> {len(final)}; dropped {sum(1 for r in orig if r['kind']=='arrow')} "
      f"old copies, added {n_copies} fresh)")

# ── placements.csv ────────────────────────────────────────────────────────────
with open(OUT_CSV, "w", newline="") as fh:
    w = csv.writer(fh)
    w.writerow(["old_id", "old_slug", "old_parent", "target_label", "target_id",
                "target_slug", "round", "score", "exact_kw", "used_synonym",
                "confidence", "moved"])
    for si, p in sorted(placement.items(),
                        key=lambda kv: (kv[1]["low_conf"], kv[1]["round"], -kv[1]["score"])):
        row = orig[si]
        tid, tslug = survivor(groups[p["gid"]])
        old = prev_parent.get(row["cid"], "(unplaced)")
        moved = "yes" if old != groups[p["gid"]]["name"] else "no"
        w.writerow([row["cid"], row["slug"], old, groups[p["gid"]]["name"],
                    tid, tslug, p["round"], p["score"], p["exact"],
                    "yes" if p["syn"] else "no",
                    "low" if p["low_conf"] else "high", moved])

# ── unplaced.csv — "anything not found, listed below" ─────────────────────────
UNPLACED_CSV = os.path.join(DESKTOP, "unplaced.csv")
with open(UNPLACED_CSV, "w", newline="") as fh:
    w = csv.writer(fh)
    w.writerow(["id", "slug", "full_category", "was_under"])
    for si in remaining:
        row = orig[si]
        w.writerow([row["cid"], row["slug"], row["full"],
                    prev_parent.get(row["cid"], "(unplaced)")])

n_low = sum(1 for p in placement.values() if p["low_conf"])
n_moved = sum(1 for si, p in placement.items()
              if prev_parent.get(orig[si]["cid"], "(unplaced)") != groups[p["gid"]]["name"])
print(f"\nPlaced {len(placement)}  ({len(placement)-n_low} high=3 exact, {n_low} low/red=2kw+synonym) | "
      f"{n_moved} re-homed | {len(remaining)} unplaced (see unplaced.csv)")
wb.save(OUT_XLSX)
print(f"Wrote {OUT_XLSX}\nWrote {OUT_CSV}\nWrote {UNPLACED_CSV}")
